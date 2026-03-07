import openpyxl
import json

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)

# Find the sheet
print("All sheets:", wb.sheetnames)

# Try to find jan 30 - feb 5 sheet
target_sheet = None
for name in wb.sheetnames:
    lower = name.lower()
    if 'jan' in lower and '30' in lower:
        target_sheet = name
        break
    if '30' in lower and '5' in lower:
        target_sheet = name
        break

if not target_sheet:
    # Show all sheets and sample data
    for name in wb.sheetnames:
        ws = wb[name]
        print(f"\nSheet: '{name}' (rows={ws.max_row}, cols={ws.max_column})")
        for row in range(1, min(5, ws.max_row+1)):
            vals = [ws.cell(row=row, column=c).value for c in range(1, min(20, ws.max_column+1))]
            vals_str = [str(v)[:30] if v else '' for v in vals]
            print(f"  Row {row}: {vals_str}")
else:
    print(f"\nFound sheet: '{target_sheet}'")
    ws = wb[target_sheet]
    
    # Scan structure - check first few rows across all columns
    print(f"\nSheet dimensions: {ws.max_row} rows x {ws.max_column} cols")
    for row in range(1, min(8, ws.max_row+1)):
        vals = []
        for c in range(1, min(20, ws.max_column+1)):
            v = ws.cell(row=row, column=c).value
            if v:
                col_letter = openpyxl.utils.get_column_letter(c)
                vals.append(f"{col_letter}={v}")
        print(f"  Row {row}: {vals}")
    
    # Find where names are by checking columns B, C, D for row 4-6
    print("\nLooking for name column:")
    for col in ['B', 'C', 'D', 'E']:
        for row in range(3, 8):
            v = ws[f'{col}{row}'].value
            if v and isinstance(v, str) and len(v) > 3:
                print(f"  {col}{row} = '{v}'")
