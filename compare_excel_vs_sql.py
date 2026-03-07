import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (1).xlsx', data_only=True)
ws = wb['feb 6-12']

print("=" * 80)
print("COMPARING EXCEL OT HOURS WITH SQL SCRIPT")
print("=" * 80)

# Extract all workers from Excel
workers_data = []
for row in range(8, 69):  # Q8:Q68
    name = ws[f'B{row}'].value
    if not name or name == 'COMMISION SCHEDULE':
        continue
    
    days = ws[f'K{row}'].value
    ot_hours = ws[f'L{row}'].value or 0
    ot_pay = ws[f'M{row}'].value or 0
    daily_rate = ws[f'N{row}'].value
    total = ws[f'Q{row}'].value
    
    workers_data.append({
        'row': row,
        'name': name,
        'days': days,
        'ot_hours': ot_hours,
        'ot_pay': ot_pay,
        'daily_rate': daily_rate,
        'total': total
    })

print(f"\nTotal workers in Q8:Q68: {len(workers_data)}")
print(f"Sum of all totals: {sum(w['total'] for w in workers_data if w['total'])}")

# Calculate what the total should be
calculated_total = 0
for w in workers_data:
    if w['days'] and w['daily_rate']:
        base_pay = w['days'] * w['daily_rate']
        ot_pay = w['ot_pay'] or 0
        worker_total = base_pay + ot_pay
        calculated_total += worker_total

print(f"Calculated total: {calculated_total}")

# Show workers with OT
print("\nWorkers with OT hours:")
ot_workers = [w for w in workers_data if w['ot_hours'] > 0]
print(f"Count: {len(ot_workers)}")

total_ot_pay = sum(w['ot_pay'] for w in ot_workers)
print(f"Total OT Pay: {total_ot_pay}")

# Show a few examples
print("\nFirst 5 workers with OT:")
for w in ot_workers[:5]:
    print(f"{w['name']}: {w['days']} days, {w['ot_hours']} OT hrs, ₱{w['ot_pay']} OT pay, Total: ₱{w['total']}")
