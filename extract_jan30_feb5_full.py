import openpyxl
import json

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)
ws = wb['JAN 30-FEB 5']

print("=" * 80)
print("EXTRACTING JAN 30 - FEB 5 DATA")
print("=" * 80)

# Structure: C=name, K=days, L=OT_hours, M=OT_pay, N=rate, O=SSS, P=subtotal, Q=total
# Data starts at row 4, header at row 7 (but data also before header)
workers = []

for row in range(4, 80):
    name = ws[f'C{row}'].value
    if not name or not isinstance(name, str):
        continue
    if any(x in name for x in ['COMMISION', 'Name:', 'TOTAL', 'total', 'Prepared']):
        continue
    
    days = ws[f'K{row}'].value
    ot_hours = ws[f'L{row}'].value or 0
    ot_pay = ws[f'M{row}'].value or 0
    daily_rate = ws[f'N{row}'].value
    sss = ws[f'O{row}'].value or 0
    subtotal = ws[f'P{row}'].value or 0
    total = ws[f'Q{row}'].value
    
    if not days or not daily_rate or total is None:
        continue
    
    # Calculate bonus: subtotal = (days*rate) + ot_pay + bonus
    expected_base = (days * daily_rate) + ot_pay
    bonus = round(subtotal - expected_base) if subtotal > expected_base + 0.5 else 0
    
    workers.append({
        'name': name.strip(),
        'days': int(days),
        'ot_hours': round(float(ot_hours), 1) if ot_hours else 0,
        'ot_pay': round(float(ot_pay), 2) if ot_pay else 0,
        'daily_rate': round(float(daily_rate)),
        'bonus': round(float(bonus)),
        'sss': round(float(sss)),
        'subtotal': round(float(subtotal), 2),
        'total': round(float(total), 2)
    })

subtotal_sum = sum(w['subtotal'] for w in workers)
sss_sum = sum(w['sss'] for w in workers)
total_sum = sum(w['total'] for w in workers)

print(f"Workers: {len(workers)}")
print(f"Subtotal (before SSS): {subtotal_sum:,.2f}")
print(f"Total SSS: {sss_sum:,.2f}")
print(f"Grand Total: {total_sum:,.2f}")

# Show first 5 workers
for i, w in enumerate(workers[:5], 1):
    print(f"  {i}. {w['name']}: {w['days']}d, {w['ot_hours']}OT, rate={w['daily_rate']}, bonus={w['bonus']}, sss={w['sss']}, total={w['total']}")

# Check for total rows in Excel
for row in range(70, 82):
    for col in ['P', 'Q']:
        v = ws[f'{col}{row}'].value
        if v and isinstance(v, (int, float)) and v > 10000:
            print(f"  Excel {col}{row} = {v:,.2f}")

# Save to JSON (same format as other periods)
period_data = {
    'jan_30_feb_5': {
        'start': '2026-01-30',
        'end': '2026-02-05',
        'workers': workers,
        'subtotal': subtotal_sum,
        'total_sss': sss_sum,
        'grand_total': total_sum
    }
}

with open('jan30_feb5_data.json', 'w', encoding='utf-8') as f:
    json.dump(period_data, f, ensure_ascii=False, indent=2)

print(f"\nSaved to jan30_feb5_data.json")
