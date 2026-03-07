import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)
ws = wb['FEB 13-19']

print("=" * 80)
print("FINDING NAMES IN FEB 13-19 SHEET")
print("=" * 80)

# Check all columns for row 5 (first data row)
print("\nRow 5 (first worker) - all columns:")
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
    val = ws[f'{col_letter}5'].value
    if val is not None:
        print(f"  {col_letter}5 = {val}")

print("\nRow 6 (second worker) - all columns:")
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
    val = ws[f'{col_letter}6'].value
    if val is not None:
        print(f"  {col_letter}6 = {val}")

print("\nRow 9 (third worker after header) - all columns:")
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
    val = ws[f'{col_letter}9'].value
    if val is not None:
        print(f"  {col_letter}9 = {val}")

# Compare with feb 6-12 structure
print("\n" + "=" * 80)
print("COMPARING WITH FEB 6-12 STRUCTURE")
print("=" * 80)

ws_feb6 = wb['feb 6-12']
print("\nFeb 6-12 Row 4 (first worker) - all columns:")
for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S']:
    val = ws_feb6[f'{col_letter}4'].value
    if val is not None:
        print(f"  {col_letter}4 = {val}")
