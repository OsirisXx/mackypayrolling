import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)
ws = wb['feb 6-12']

print("=" * 80)
print("EXTRACTING FEB 6-12 DATA")
print("=" * 80)

workers_data = []

# Based on scan: B=employee_id, C=name, K=days, L=OT_hours, M=OT_pay, N=daily_rate, O=SSS, P=subtotal, Q=total
for row in range(4, 80):
    employee_id = ws[f'B{row}'].value
    name = ws[f'C{row}'].value
    
    if not name or not isinstance(name, str):
        continue
    
    if 'COMMISION' in name or 'Name:' in name or 'TOTAL' in name:
        continue
    
    days = ws[f'K{row}'].value
    ot_hours = ws[f'L{row}'].value or 0
    ot_pay = ws[f'M{row}'].value or 0
    daily_rate = ws[f'N{row}'].value
    sss = ws[f'O{row}'].value or 0
    subtotal = ws[f'P{row}'].value or 0
    total = ws[f'Q{row}'].value
    
    if not days or not daily_rate or not total:
        continue
    
    # Extract bonus from formula or calculate
    bonus = 0
    if subtotal and total:
        # bonus is included in subtotal but not in total calculation
        # subtotal = days*rate + ot_pay + bonus
        # total = subtotal - sss
        expected_base = (days * daily_rate) + ot_pay
        if subtotal > expected_base:
            bonus = subtotal - expected_base
    
    workers_data.append({
        'name': name.strip(),
        'days': int(days) if days else 0,
        'ot_hours': int(ot_hours) if ot_hours else 0,
        'daily_rate': int(daily_rate) if daily_rate else 0,
        'bonus': int(bonus) if bonus else 0,
        'sss': int(sss) if sss else 0,
        'total': float(total) if total else 0
    })

print(f"\nTotal workers: {len(workers_data)}")
print(f"Sum of totals: {sum(w['total'] for w in workers_data):,.2f}")

# Show first 10 workers
print("\nFirst 10 workers:")
for i, w in enumerate(workers_data[:10], 1):
    print(f"{i}. {w['name']}: {w['days']} days, {w['ot_hours']} OT, Rate={w['daily_rate']}, Bonus={w['bonus']}, SSS={w['sss']}, Total={w['total']:,.2f}")

# Save to file
with open('feb_6_12_data.txt', 'w', encoding='utf-8') as f:
    f.write("name|days|ot_hours|daily_rate|bonus|sss|total\n")
    for w in workers_data:
        f.write(f"{w['name']}|{w['days']}|{w['ot_hours']}|{w['daily_rate']}|{w['bonus']}|{w['sss']}|{w['total']}\n")

print(f"\nData saved to feb_6_12_data.txt")
print(f"Grand Total: {sum(w['total'] for w in workers_data):,.2f}")
