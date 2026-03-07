import openpyxl

# Load the workbook
wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=False)

# Get the Feb 13-19 sheet
ws = wb['FEB 13-19']

print("=" * 60)
print("Checking Bacol, Vivian (Row 4):")
print("=" * 60)
print(f"  Days (D4): {ws['D4'].value}")
print(f"  OT (E4): {ws['E4'].value}")
print(f"  Rate (F4): {ws['F4'].value}")
print(f"  Bonus (G4): {ws['G4'].value}")
print(f"  Subtotal (H4): {ws['H4'].value}")
print(f"  Total (I4): {ws['I4'].value}")
print()

# Calculate what it should be
days = ws['D4'].value or 0
ot = ws['E4'].value or 0
rate = ws['F4'].value or 0
bonus = ws['G4'].value or 0

basePay = days * rate
otPay = ot * (rate / 8)
calculated = basePay + otPay + bonus

print(f"  Calculated: {days} * {rate} + {ot} * ({rate}/8) + {bonus}")
print(f"  = {basePay} + {otPay} + {bonus} = {calculated}")
print(f"  Excel shows: {ws['H4'].value}")
print(f"  Difference: {ws['H4'].value - calculated if ws['H4'].value else 'N/A'}")
print()

print("=" * 60)
print("Checking Ylagan, Robert (Row 5):")
print("=" * 60)
print(f"  Days (D5): {ws['D5'].value}")
print(f"  OT (E5): {ws['E5'].value}")
print(f"  Rate (F5): {ws['F5'].value}")
print(f"  Bonus (G5): {ws['G5'].value}")
print(f"  Subtotal (H5): {ws['H5'].value}")
print(f"  Total (I5): {ws['I5'].value}")
print()

# Calculate what it should be
days2 = ws['D5'].value or 0
ot2 = ws['E5'].value or 0
rate2 = ws['F5'].value or 0
bonus2 = ws['G5'].value or 0

basePay2 = days2 * rate2
otPay2 = ot2 * (rate2 / 8)
calculated2 = basePay2 + otPay2 + bonus2

print(f"  Calculated: {days2} * {rate2} + {ot2} * ({rate2}/8) + {bonus2}")
print(f"  = {basePay2} + {otPay2} + {bonus2} = {calculated2}")
print(f"  Excel shows: {ws['H5'].value}")
print(f"  Difference: {ws['H5'].value - calculated2 if ws['H5'].value else 'N/A'}")

# Load with formulas
wb_formulas = openpyxl.load_workbook('PLANTA-DON2-2026 (1).xlsx', data_only=False)
wb_values = openpyxl.load_workbook('PLANTA-DON2-2026 (1).xlsx', data_only=True)

ws_f = wb_formulas['feb 6-12']
ws_v = wb_values['feb 6-12']

print("=" * 80)
print("CHECKING FOR HIDDEN FORMULAS/BONUSES")
print("=" * 80)

# Check if there are any special formulas in the total column
print("\nChecking formulas for workers Q8-Q20:")
for row in range(8, 21):
    name = ws_v[f'B{row}'].value
    if not name or name == 'COMMISION SCHEDULE':
        continue
    
    formula_P = ws_f[f'P{row}'].value  # Subtotal
    formula_Q = ws_f[f'Q{row}'].value  # Total
    value_P = ws_v[f'P{row}'].value
    value_Q = ws_v[f'Q{row}'].value
    
    print(f"\nRow {row}: {name}")
    print(f"  P{row} formula: {formula_P}")
    print(f"  P{row} value: {value_P}")
    print(f"  Q{row} formula: {formula_Q}")
    print(f"  Q{row} value: {value_Q}")

# Check for any bonus columns
print("\n" + "=" * 80)
print("Checking columns O, P, Q for special values:")
print("=" * 80)

for row in [4, 5, 8, 9, 10]:
    name = ws_v[f'B{row}'].value
    if name:
        print(f"\nRow {row}: {name}")
        for col in ['O', 'P', 'Q']:
            formula = ws_f[f'{col}{row}'].value
            value = ws_v[f'{col}{row}'].value
            print(f"  {col}{row}: formula={formula}, value={value}")
