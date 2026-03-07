import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (1).xlsx', data_only=True)
ws = wb['feb 6-12']

print("=" * 80)
print("FEB 6-12 ANALYSIS (NEW FILE)")
print("=" * 80)

print(f"\nGrand Total (Q70): {ws['Q70'].value}")

print("\nFirst 10 workers:")
for row in range(4, 14):
    name = ws[f'B{row}'].value
    days = ws[f'K{row}'].value
    ot_hours = ws[f'L{row}'].value
    ot_pay = ws[f'M{row}'].value
    daily_rate = ws[f'N{row}'].value
    total = ws[f'Q{row}'].value
    
    if name:
        print(f"Row {row}: {name}")
        print(f"  Days: {days}, OT Hours: {ot_hours}, OT Pay: {ot_pay}")
        print(f"  Daily Rate: {daily_rate}, Total: {total}")
        print()

# Check sum of all workers
print("\nCalculating totals:")
workers_sum = sum(ws[f'Q{i}'].value for i in range(8, 69) if ws[f'Q{i}'].value)
print(f"Sum Q8:Q68: {workers_sum}")
print(f"Q4 (Bacol): {ws['Q4'].value}")
print(f"Q5 (Ylagan): {ws['Q5'].value}")
print(f"Grand Total: {workers_sum + ws['Q4'].value + ws['Q5'].value}")
print(f"Excel Q70: {ws['Q70'].value}")
