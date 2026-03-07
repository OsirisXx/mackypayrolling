import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)
ws = wb['feb 6-12']

print("=" * 80)
print("FEB 6-12 ANALYSIS - NEW FILE (2)")
print("=" * 80)

print(f"\nGrand Total (Q70): {ws['Q70'].value}")

# Check all workers and their totals
print("\nAll workers with their data:")
total_sum = 0
worker_count = 0

for row in range(4, 69):
    name = ws[f'B{row}'].value
    if not name or name == 'COMMISION SCHEDULE':
        continue
    
    days = ws[f'K{row}'].value
    ot_hours = ws[f'L{row}'].value
    ot_pay = ws[f'M{row}'].value or 0
    daily_rate = ws[f'N{row}'].value
    total = ws[f'Q{row}'].value
    
    if total:
        total_sum += total
        worker_count += 1
        
        # Show first 10 and any with differences
        if worker_count <= 10:
            print(f"{name}: Days={days}, OT hrs={ot_hours}, OT pay={ot_pay}, Rate={daily_rate}, Total={total}")

print(f"\nTotal workers: {worker_count}")
print(f"Sum of all worker totals: {total_sum}")
print(f"Excel Q70 Grand Total: {ws['Q70'].value}")
print(f"Difference: {ws['Q70'].value - total_sum if ws['Q70'].value else 'N/A'}")

# Check for any formulas with bonuses
print("\n" + "=" * 80)
print("Checking for hidden bonuses in formulas...")
print("=" * 80)

wb_formulas = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=False)
ws_f = wb_formulas['feb 6-12']

bonus_count = 0
for row in range(4, 69):
    name = ws[f'B{row}'].value
    if not name or name == 'COMMISION SCHEDULE':
        continue
    
    formula_P = ws_f[f'P{row}'].value
    if formula_P and isinstance(formula_P, str) and '+' in formula_P:
        # Check if it's a bonus (not just addition of cells)
        if any(char.isdigit() for char in formula_P.split('+')[-1]):
            bonus_count += 1
            if bonus_count <= 5:
                print(f"Row {row} ({name}): {formula_P}")

print(f"\nTotal workers with bonuses in formulas: {bonus_count}")
