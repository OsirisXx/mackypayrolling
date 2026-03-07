import openpyxl

# Load both formula and data versions
wb_formulas = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=False)
wb_values = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)

print("=" * 80)
print("ANALYZING ALL SHEETS FOR BACOL AND YLAGAN PATTERNS")
print("=" * 80)

for sheet_name in wb_formulas.sheetnames:
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print("=" * 80)
    
    ws_f = wb_formulas[sheet_name]
    ws_v = wb_values[sheet_name]
    
    # Find Bacol and Ylagan rows by scanning first 10 rows
    bacol_row = None
    ylagan_row = None
    
    for row_num in range(1, 15):
        for col_num in range(1, 10):
            cell_val = ws_f.cell(row=row_num, column=col_num).value
            if cell_val and isinstance(cell_val, str):
                cell_lower = cell_val.lower()
                if 'bacol' in cell_lower and 'vivian' in cell_lower:
                    bacol_row = row_num
                if 'ylagan' in cell_lower and 'robert' in cell_lower:
                    ylagan_row = row_num
    
    if bacol_row:
        print(f"\n--- BACOL, VIVIAN (Row {bacol_row}) ---")
        row_data = list(ws_f[bacol_row])
        row_vals = list(ws_v[bacol_row])
        
        # Print all non-empty cells with their column letters
        for i, (cell_f, cell_v) in enumerate(zip(row_data, row_vals)):
            col_letter = openpyxl.utils.get_column_letter(i + 1)
            if cell_f.value is not None or cell_v.value is not None:
                formula = cell_f.value
                value = cell_v.value
                if formula != value:
                    print(f"  {col_letter}{bacol_row}: Formula='{formula}', Value={value}")
                else:
                    print(f"  {col_letter}{bacol_row}: {value}")
    else:
        print("  Bacol not found in this sheet")
    
    if ylagan_row:
        print(f"\n--- YLAGAN, ROBERT (Row {ylagan_row}) ---")
        row_data = list(ws_f[ylagan_row])
        row_vals = list(ws_v[ylagan_row])
        
        # Print all non-empty cells with their column letters
        for i, (cell_f, cell_v) in enumerate(zip(row_data, row_vals)):
            col_letter = openpyxl.utils.get_column_letter(i + 1)
            if cell_f.value is not None or cell_v.value is not None:
                formula = cell_f.value
                value = cell_v.value
                if formula != value:
                    print(f"  {col_letter}{ylagan_row}: Formula='{formula}', Value={value}")
                else:
                    print(f"  {col_letter}{ylagan_row}: {value}")
    else:
        print("  Ylagan not found in this sheet")
    
    # Also check regular workers to understand the standard formula
    print(f"\n--- SAMPLE REGULAR WORKER (checking row 10) ---")
    row_data = list(ws_f[10])
    row_vals = list(ws_v[10])
    for i, (cell_f, cell_v) in enumerate(zip(row_data, row_vals)):
        col_letter = openpyxl.utils.get_column_letter(i + 1)
        if cell_f.value is not None or cell_v.value is not None:
            formula = cell_f.value
            value = cell_v.value
            if formula != value:
                print(f"  {col_letter}10: Formula='{formula}', Value={value}")
            else:
                print(f"  {col_letter}10: {value}")

print("\n" + "=" * 80)
print("ANALYSIS COMPLETE")
print("=" * 80)
