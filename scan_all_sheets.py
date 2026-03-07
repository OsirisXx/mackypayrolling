import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)

print("=" * 80)
print("SCANNING ALL SHEETS IN EXCEL FILE")
print("=" * 80)

for sheet_name in wb.sheetnames:
    print(f"\n{'=' * 80}")
    print(f"SHEET: {sheet_name}")
    print('=' * 80)
    
    ws = wb[sheet_name]
    
    # Find first non-empty row with data
    first_data_row = None
    for row in range(1, min(100, ws.max_row + 1)):
        # Check if row has substantial data (at least 3 non-empty cells)
        non_empty = sum(1 for col in range(1, min(20, ws.max_column + 1)) if ws.cell(row, col).value is not None)
        if non_empty >= 3:
            first_data_row = row
            break
    
    if first_data_row:
        print(f"First data row: {first_data_row}")
        print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
        
        # Show first 5 rows of data
        print("\nFirst 5 rows:")
        for row in range(first_data_row, min(first_data_row + 5, ws.max_row + 1)):
            row_data = []
            for col in ['A', 'B', 'C', 'K', 'L', 'M', 'N', 'O', 'P', 'Q']:
                val = ws[f'{col}{row}'].value
                if val is not None:
                    row_data.append(f"{col}={val}")
            if row_data:
                print(f"  Row {row}: {', '.join(row_data)}")
        
        # Check for total row
        for row in range(ws.max_row - 5, ws.max_row + 1):
            q_val = ws[f'Q{row}'].value
            if q_val and isinstance(q_val, (int, float)) and q_val > 100000:
                print(f"\nGrand Total at Q{row}: {q_val:,.2f}")
                break
    else:
        print("No data found in this sheet")

print("\n" + "=" * 80)
