import openpyxl
import re

wb_values = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)
wb_formulas = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=False)

print("Available sheets:")
for sheet_name in wb_values.sheetnames:
    print(f"  - {sheet_name}")

# Try to find the Feb 13-19 sheet
sheet_name = None
for name in wb_values.sheetnames:
    if '13' in name and '19' in name:
        sheet_name = name
        print(f"\nUsing sheet: {sheet_name}")
        break

if not sheet_name:
    print("\nNo Feb 13-19 sheet found. Exiting.")
    exit(1)

ws_v = wb_values[sheet_name]
ws_f = wb_formulas[sheet_name]

print("=" * 80)
print("EXTRACTING FEB 13-19 DATA")
print("=" * 80)

# Check if sheet has any data at all
print("\nChecking if sheet has any data...")
print(f"Max row: {ws_v.max_row}")
print(f"Max column: {ws_v.max_column}")

# Scan all cells to find any non-empty cell
found_data = False
for row in range(1, min(100, ws_v.max_row + 1)):
    for col in range(1, min(30, ws_v.max_column + 1)):
        cell = ws_v.cell(row, col)
        if cell.value is not None:
            print(f"Found data at row {row}, col {col}: {cell.value}")
            found_data = True
            if not found_data:
                break
    if found_data:
        break

if not found_data:
    print("\n⚠️ FEB 13-19 sheet is completely empty!")
    print("Please fill in the Feb 13-19 data in the Excel file first.")
    exit(1)

# Find where data starts by looking for any worker name
print("\nSearching for worker data...")
data_start_row = None
for row in range(1, 100):
    cell_val = ws_v[f'B{row}'].value
    if cell_val and isinstance(cell_val, str):
        # Look for common worker names or patterns
        if any(name in cell_val for name in ['Bacol', 'Vivian', 'Ylagan', 'Robert', 'Abaday', 'Emelyn']):
            data_start_row = row
            print(f"✓ Data starts at row {row}: {cell_val}")
            break
        # Also check if it looks like a name (has comma or multiple words)
        if ',' in cell_val or (len(cell_val.split()) >= 2 and cell_val[0].isupper()):
            print(f"  Possible name at row {row}: {cell_val}")

if not data_start_row:
    print("\n⚠️ Couldn't find worker data. Showing all non-empty cells in column B:")
    for row in range(1, 50):
        val = ws_v[f'B{row}'].value
        if val:
            print(f"  B{row}: {val}")
    exit(1)

print(f"\nGrand Total (Q70): {ws_v['Q70'].value}")

workers_data = []

for row in range(data_start_row, 70):
    employee_id = ws_v[f'A{row}'].value
    name = ws_v[f'B{row}'].value
    
    if not name or name == 'COMMISION SCHEDULE':
        continue
    
    days = ws_v[f'K{row}'].value
    ot_hours = ws_v[f'L{row}'].value or 0
    ot_pay = ws_v[f'M{row}'].value or 0
    daily_rate = ws_v[f'N{row}'].value
    sss = ws_v[f'O{row}'].value or 0
    total = ws_v[f'Q{row}'].value
    
    # Check formula for bonus
    formula_P = ws_f[f'P{row}'].value
    bonus = 0
    if formula_P and isinstance(formula_P, str):
        bonus_match = re.search(r'\+(\d+)$', formula_P)
        if bonus_match:
            bonus = int(bonus_match.group(1))
    
    workers_data.append({
        'row': row,
        'employee_id': employee_id,
        'name': name,
        'days': days,
        'ot_hours': ot_hours,
        'ot_pay': ot_pay,
        'daily_rate': daily_rate,
        'bonus': bonus,
        'sss': sss,
        'total': total
    })

print(f"\nTotal workers: {len(workers_data)}")
print(f"Sum of all totals: ₱{sum(w['total'] for w in workers_data if w['total']):,.2f}")

# Show workers with bonuses or SSS
print("\n" + "=" * 80)
print("WORKERS WITH BONUSES OR SSS:")
print("=" * 80)

workers_with_adjustments = [w for w in workers_data if w['bonus'] > 0 or w['sss'] > 0]
print(f"Found {len(workers_with_adjustments)} workers with bonuses/SSS\n")

for w in workers_with_adjustments:
    print(f"{w['name']}: Bonus=+{w['bonus']}, SSS=-{w['sss']}")

# Show first 10 workers for verification
print("\n" + "=" * 80)
print("FIRST 10 WORKERS:")
print("=" * 80)

for i, w in enumerate(workers_data[:10], 1):
    print(f"{i}. {w['name']}: {w['days']} days, {w['ot_hours']} OT hrs, ₱{w['total']}")

print("\n" + "=" * 80)
print(f"Total bonuses: ₱{sum(w['bonus'] for w in workers_data)}")
print(f"Total SSS: ₱{sum(w['sss'] for w in workers_data)}")
print(f"Grand Total: ₱{sum(w['total'] for w in workers_data if w['total']):,.2f}")
print("=" * 80)
