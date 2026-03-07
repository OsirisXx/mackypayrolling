import openpyxl
import json

wb = openpyxl.load_workbook(r'd:\Commission\macky-attendance\PLANTA-DON2-2026 (2).xlsx', data_only=False)
ws = wb['FEB 13-19']

# Also load with data_only to get computed values
wb_data = openpyxl.load_workbook(r'd:\Commission\macky-attendance\PLANTA-DON2-2026 (2).xlsx', data_only=True)
ws_data = wb_data['FEB 13-19']

workers = []
for row_num in range(5, ws.max_row + 1):
    name = ws.cell(row=row_num, column=4).value  # D column
    if not name or name == 'Name:  ':
        continue
    
    # Get raw values and formulas
    worker = {
        'row': row_num,
        'name': name,
        'days_detail': [],
        'L_days': ws.cell(row=row_num, column=12).value,  # L - total days
        'M_ot': ws.cell(row=row_num, column=13).value,  # M - OT hours
        'N_total_ot_formula': ws.cell(row=row_num, column=14).value,  # N - total OT (formula)
        'N_total_ot_value': ws_data.cell(row=row_num, column=14).value,  # N - total OT (computed)
        'O_daily_rate': ws.cell(row=row_num, column=15).value,  # O - Daily Rate
        'P_bonus_formula': ws.cell(row=row_num, column=16).value,  # P - bonus/deduction (formula)
        'P_bonus_value': ws_data.cell(row=row_num, column=16).value,  # P - bonus/deduction (computed)
        'Q_total_formula': ws.cell(row=row_num, column=17).value,  # Q - total (formula)
        'Q_total_value': ws_data.cell(row=row_num, column=17).value,  # Q - total (computed)
        'R_net_formula': ws.cell(row=row_num, column=18).value,  # R - net (formula)
        'R_net_value': ws_data.cell(row=row_num, column=18).value,  # R - net (computed)
    }
    
    # Get daily values (E through K)
    for col in range(5, 12):
        val = ws.cell(row=row_num, column=col).value
        worker['days_detail'].append(val)
    
    workers.append(worker)

# Print all workers
for w in workers:
    print(f"\n=== {w['name']} (row {w['row']}) ===")
    print(f"  Days detail: {w['days_detail']}")
    print(f"  L(days): {w['L_days']}, M(OT): {w['M_ot']}")
    print(f"  N(total OT): formula={w['N_total_ot_formula']}, value={w['N_total_ot_value']}")
    print(f"  O(daily rate): {w['O_daily_rate']}")
    print(f"  P(bonus/ded): formula={w['P_bonus_formula']}, value={w['P_bonus_value']}")
    print(f"  Q(total): formula={w['Q_total_formula']}, value={w['Q_total_value']}")
    print(f"  R(net): formula={w['R_net_formula']}, value={w['R_net_value']}")

# Identify unique Q formulas
print("\n\n=== UNIQUE Q (TOTAL) FORMULAS ===")
formulas = {}
for w in workers:
    q = w['Q_total_formula']
    if q not in formulas:
        formulas[q] = []
    formulas[q].append(w['name'])

for f, names in formulas.items():
    print(f"\nFormula: {f}")
    print(f"  Workers ({len(names)}): {', '.join(names[:5])}{'...' if len(names) > 5 else ''}")

# Identify unique P formulas (bonus patterns)
print("\n\n=== UNIQUE P (BONUS/DEDUCTION) FORMULAS ===")
p_formulas = {}
for w in workers:
    p = w['P_bonus_formula']
    if p not in p_formulas:
        p_formulas[p] = []
    p_formulas[p].append(w['name'])

for f, names in p_formulas.items():
    print(f"\nP value/formula: {f}")
    print(f"  Workers ({len(names)}): {', '.join(names[:5])}{'...' if len(names) > 5 else ''}")
