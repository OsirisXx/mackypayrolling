import openpyxl
import json

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)

sheets = {
    'jan_30_feb_5': {'sheet': 'JAN 30-FEB 5', 'start': '2026-01-30', 'end': '2026-02-05'},
    'feb_6_12': {'sheet': 'feb 6-12', 'start': '2026-02-06', 'end': '2026-02-12'},
    'feb_13_19': {'sheet': 'FEB 13-19', 'start': '2026-02-13', 'end': '2026-02-19'},
}

# For feb 6-12 and jan 30-feb 5: C=name, K=days, L=OT_hours, M=OT_pay, N=rate, O=bonus/SSS, P=subtotal, Q=total
# For feb 13-19: D=name, L=days, M=OT_hours, N=OT_pay, O=rate, P=bonus/SSS, Q=subtotal, R=total

all_data = {}

for key, info in sheets.items():
    ws = wb[info['sheet']]
    print(f"\n{'='*80}")
    print(f"Sheet: {info['sheet']} ({key})")
    
    # Detect structure by checking row 7 header
    is_feb_13_19 = (key == 'feb_13_19')
    
    if is_feb_13_19:
        name_col = 'D'
        days_col = 'L'
        ot_hours_col = 'M'
        ot_pay_col = 'N'
        rate_col = 'O'
        sss_col = 'P'
        subtotal_col = 'Q'
        total_col = 'R'
    else:
        name_col = 'C'
        days_col = 'K'
        ot_hours_col = 'L'
        ot_pay_col = 'M'
        rate_col = 'N'
        sss_col = 'O'
        subtotal_col = 'P'
        total_col = 'Q'
    
    workers = []
    skipped = []
    
    for row in range(4, 85):
        name = ws[f'{name_col}{row}'].value
        if not name or not isinstance(name, str):
            continue
        name = name.strip()
        if any(x in name.lower() for x in ['commis', 'name:', 'total', 'prepared', 'powered', 'period', 'macrock', 'signature', 'schedule']):
            continue
        if len(name) < 2:
            continue
        
        days = ws[f'{days_col}{row}'].value or 0
        ot_hours = ws[f'{ot_hours_col}{row}'].value or 0
        ot_pay = ws[f'{ot_pay_col}{row}'].value or 0
        daily_rate = ws[f'{rate_col}{row}'].value or 0
        sss_raw = ws[f'{sss_col}{row}'].value or 0
        subtotal = ws[f'{subtotal_col}{row}'].value or 0
        total = ws[f'{total_col}{row}'].value or 0
        
        # Try to get rate even for 0-day workers
        if daily_rate == 0:
            daily_rate = 400  # default rate
        
        try:
            days = int(float(days)) if days else 0
        except:
            days = 0
        
        try:
            ot_pay = round(float(ot_pay), 2) if ot_pay else 0
        except:
            ot_pay = 0
            
        try:
            daily_rate = round(float(daily_rate))
        except:
            daily_rate = 400
            
        try:
            sss = round(float(sss_raw)) if sss_raw else 0
        except:
            sss = 0
            
        try:
            subtotal = round(float(subtotal), 2) if subtotal else 0
        except:
            subtotal = 0
            
        try:
            total = round(float(total), 2) if total else 0
        except:
            total = 0
        
        # Calculate bonus
        expected_base = (days * daily_rate) + ot_pay
        bonus = round(subtotal - expected_base) if subtotal > expected_base + 0.5 else 0
        
        workers.append({
            'name': name,
            'days': days,
            'ot_hours': round(float(ot_hours), 1) if ot_hours else 0,
            'ot_pay': ot_pay,
            'daily_rate': daily_rate,
            'bonus': bonus,
            'sss': sss,
            'subtotal': subtotal,
            'total': total
        })
    
    subtotal_sum = sum(w['subtotal'] for w in workers)
    sss_sum = sum(w['sss'] for w in workers)
    total_sum = sum(w['total'] for w in workers)
    
    workers_with_days = [w for w in workers if w['days'] > 0]
    workers_without_days = [w for w in workers if w['days'] == 0]
    
    print(f"Total workers: {len(workers)} ({len(workers_with_days)} with days, {len(workers_without_days)} with 0 days)")
    print(f"Subtotal: {subtotal_sum:,.2f} | SSS: {sss_sum:,.2f} | Grand Total: {total_sum:,.2f}")
    
    if workers_without_days:
        print(f"\nWorkers with 0 days:")
        for w in workers_without_days:
            print(f"  {w['name']}: days={w['days']}, rate={w['daily_rate']}, total={w['total']}")
    
    all_data[key] = {
        'start': info['start'],
        'end': info['end'],
        'workers': workers,
        'subtotal': subtotal_sum,
        'total_sss': sss_sum,
        'grand_total': total_sum
    }

# Save
with open('payroll_data_all.json', 'w', encoding='utf-8') as f:
    json.dump(all_data, f, ensure_ascii=False, indent=2)

print(f"\n\nSaved all data to payroll_data_all.json")
