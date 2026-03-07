import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026.xlsx', data_only=True)
ws = wb['feb 6-12']

print("Worker OT Hours from Excel (Feb 6-12):")
print("=" * 60)

# Start from row 4 (Bacol) to row 68
for row in range(4, 69):
    name = ws[f'B{row}'].value
    if not name:
        continue
    
    days = ws[f'K{row}'].value
    ot_hours = ws[f'L{row}'].value
    ot_pay = ws[f'M{row}'].value
    daily_rate = ws[f'N{row}'].value
    total = ws[f'Q{row}'].value
    
    if ot_hours:
        print(f"Row {row}: {name}")
        print(f"  Days: {days}, OT Hours: {ot_hours}, OT Pay: {ot_pay}, Daily Rate: {daily_rate}")
        print(f"  Total: {total}")
        print()
