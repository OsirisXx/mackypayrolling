import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)
ws = wb['feb 6-12']

print("=" * 80)
print("FINDING THE ₱350 DIFFERENCE")
print("=" * 80)

# Get all Excel workers with their totals
excel_workers = []
for row in range(4, 69):
    name = ws[f'B{row}'].value
    if name and name != 'COMMISION SCHEDULE':
        total = ws[f'Q{row}'].value
        if total:
            excel_workers.append({
                'row': row,
                'name': str(name),
                'total': total
            })

print(f"\nExcel has {len(excel_workers)} workers")
print(f"Excel total: ₱{sum(w['total'] for w in excel_workers):,.2f}")

# System shows 62 workers with total 169,836.25
# Excel shows 61 workers with total 170,186.25
# Difference: 350

# Check if there's a worker with total around 350 that might be duplicated or missing
print("\n" + "=" * 80)
print("Looking for workers with totals around ₱350 or differences:")
print("=" * 80)

for w in excel_workers:
    if 300 <= w['total'] <= 400:
        print(f"Row {w['row']}: {w['name']} = ₱{w['total']}")

# List all workers to compare
print("\n" + "=" * 80)
print("ALL EXCEL WORKERS:")
print("=" * 80)

for i, w in enumerate(excel_workers, 1):
    print(f"{i}. {w['name']}: ₱{w['total']}")
