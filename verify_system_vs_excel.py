import openpyxl

wb = openpyxl.load_workbook('PLANTA-DON2-2026 (2).xlsx', data_only=True)
ws = wb['feb 6-12']

print("=" * 80)
print("VERIFYING SYSTEM VS EXCEL TOTALS")
print("=" * 80)

# System totals (from screenshot)
system_workers = {
    'Bacol, Vivian': 2727.50,
    'Ylagan Robert': 3602.50,
    'Abaday Emelyn': 2100.00,
    'Abaday Ike Jun': 2550.00,
    'Absin Jimmy': 1150.00,
    'Abunda, Renjay': 3050.00,
    'Alimbog, Gabby': 2700.00,
    'Alimbog, Livy': 2400.00,
    'Alimbog,marque': 2350.00,
    'Amoncio,Jendel': 2400.00,
    'Angan, e': 2450.00,
    'Arnesx Sarinao': 4150.00,
    'Atlawan, jimboy': 2250.00,
    'Baculio Daren': 3250.00,
    'Baculio Rosito': 2450.00,
    'Balbuena Randy': 2100.00,
    'Bulak, Alvin': 2800.00,
    'Bulak Marvin': 4650.00,
    'Bulak Melvin': 2450.00,
    'Bulak Norvin': 4050.00,
    'Cabornay, Celio': 2800.00,
    'Conahan,Aaron': 2000.00,
    'Curbo Regie': 3800.00,
    'Dagalia Cruz Popoy': 2100.00,
    'Decaso Jclaid': 3350.00,
    'Espina Aida': 2550.00,
    'Felizarda,Lintacan': 1950.00,
    'Funa, Muela': 3637.50,
    'Gullan MEnard': 2400.00,
    'Mahunyag, christ': 2400.00,
    'Marjunel Angan': 3000.00,
    'Pagayon Joanrd': 2800.00,
    'Pagayon jason': 2450.00,
    'Renel Angan': 5343.75,
    'Lastimosa Ricky': 4100.00,
    'Sagansan, Dexter': 2000.00,
    'Sagansan Jerry': 3400.00,
    'Salait Ariel': 2450.00,
    'Sumayan Roland': 800.00,
    'Yake jordan': 1900.00,
    'Yake rodonio': 2000.00,
    'Yanuhon JOvil': 2200.00,
    'Yunson Alfred': 4300.00,
    'Alimbog, jhon': 3200.00,
    'Alimbog Hilbert': 1600.00,
    'Andaol Zerbi': 2750.00,
    'Dagpong crist': 2850.00,
    'Datuin James': 2750.00,
    'sagansan renie': 4250.00,
    'Pagayod Beebth': 2750.00,
    'ubanan Enel': 2200.00,
    'Yake rex': 800.00,
    'yUnson niper': 4600.00,
    'Abunda Bobby': 3600.00,
    'Lan ayan Raul': 3150.00,
    'bayantong Edgar': 2625.00,
    'Camahay Diego': 4850.00,
    'Cartahan, Kevin': 1800.00,
    'Sandalan jevin': 1800.00,
    'Lan-ayan Renemee': 2450.00,
    'Gonhay Jemmuel': 2450.00,
}

system_total = sum(system_workers.values())
print(f"\nSystem Total: ₱{system_total:,.2f}")
print(f"Excel Total (Q70): ₱{ws['Q70'].value:,.2f}")
print(f"Difference: ₱{ws['Q70'].value - system_total:,.2f}")

# Check each worker in Excel
print("\n" + "=" * 80)
print("CHECKING EACH WORKER:")
print("=" * 80)

differences = []
for row in range(4, 69):
    name = ws[f'B{row}'].value
    if not name or name == 'COMMISION SCHEDULE':
        continue
    
    excel_total = ws[f'Q{row}'].value
    if not excel_total:
        continue
    
    # Try to find matching name in system
    system_total_worker = None
    for sys_name, sys_total in system_workers.items():
        if sys_name.lower().strip() in name.lower().strip() or name.lower().strip() in sys_name.lower().strip():
            system_total_worker = sys_total
            break
    
    if system_total_worker:
        diff = excel_total - system_total_worker
        if abs(diff) > 0.01:
            differences.append({
                'name': name,
                'excel': excel_total,
                'system': system_total_worker,
                'diff': diff
            })

if differences:
    print(f"\nFound {len(differences)} workers with differences:")
    for d in differences:
        print(f"{d['name']}: Excel=₱{d['excel']}, System=₱{d['system']}, Diff=₱{d['diff']}")
else:
    print("\nAll workers match!")

# Check if there are workers in Excel not in system
print("\n" + "=" * 80)
print("WORKERS IN EXCEL BUT NOT IN SYSTEM:")
print("=" * 80)

excel_workers = []
for row in range(4, 69):
    name = ws[f'B{row}'].value
    if name and name != 'COMMISION SCHEDULE':
        total = ws[f'Q{row}'].value
        if total:
            excel_workers.append((name, total))

print(f"Excel has {len(excel_workers)} workers")
print(f"System has {len(system_workers)} workers")

if len(excel_workers) > len(system_workers):
    print(f"\nMissing {len(excel_workers) - len(system_workers)} workers in system!")
    for name, total in excel_workers:
        found = False
        for sys_name in system_workers.keys():
            if sys_name.lower().strip() in name.lower().strip() or name.lower().strip() in sys_name.lower().strip():
                found = True
                break
        if not found:
            print(f"  {name}: ₱{total}")
